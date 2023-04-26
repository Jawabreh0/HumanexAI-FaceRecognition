import cv2
import numpy as np
from mtcnn import MTCNN
from tensorflow.keras.models import load_model
from sklearn.preprocessing import Normalizer, LabelEncoder
from sklearn.svm import SVC
import openpyxl
import os 
# Load face detection model
detector = MTCNN()

# Load face recognition model
facenet_model = load_model('/home/jawabreh/Desktop/face-recognition/facenet_keras.h5')

# Load face embeddings
data = np.load('/home/jawabreh/Desktop/face-recognition/embeddings/unmasked-embeddings.npz')
trainX, trainy = data['arr_0'], data['arr_1']

# Normalize input vectors
in_encoder = Normalizer(norm='l2')
trainX = in_encoder.transform(trainX)

# Label encode targets
out_encoder = LabelEncoder()
out_encoder.fit(trainy)
trainy = out_encoder.transform(trainy)

# Define the classes
class_names = out_encoder.classes_
class_names = np.append(class_names, 'unknown')

# Train SVM classifier
model = SVC(kernel='linear', probability=True)
model.fit(trainX, trainy)

# Define function to extract face embeddings
def extract_face_embeddings(image):
    # Detect faces in the image
    faces = detector.detect_faces(image)
    if not faces:
        return None
    # Extract the first face only
    x1, y1, width, height = faces[0]['box']
    x2, y2 = x1 + width, y1 + height
    face = image[y1:y2, x1:x2]
    # Resize face to the size required by facenet model
    face = cv2.resize(face, (160, 160))
    # Preprocess the face for facenet model
    face = face.astype('float32')
    mean, std = face.mean(), face.std()
    face = (face - mean) / std
    face = np.expand_dims(face, axis=0)
    # Generate embeddings using facenet model
    embeddings = facenet_model.predict(face)
    return embeddings[0]

# Define function to identify the identity of an input image
def identify_person(image):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    prediction = model.predict(embeddings)
    confidence = model.predict_proba(embeddings)[0][prediction] * 100
    prediction = out_encoder.inverse_transform(prediction)
    return prediction[0].item(), confidence

# Define function to identify the identity and confidence of an input image
def identify_person_with_unknown(image, threshold=0.9):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    predictions = model.predict_proba(embeddings)[0]
    max_idx = np.argmax(predictions)
    if predictions[max_idx] >= threshold:
        prediction = out_encoder.inverse_transform([max_idx])
        confidence = predictions[max_idx]
        return prediction[0].item(), confidence
    else:
        return None, None

# Define path to directory containing images
image_dir = '/home/jawabreh/Desktop/face-recognition/data/val/unknown/'


# Create an empty list to store the results
results_list = []

# Iterate through each image in the directory
for filename in os.listdir(image_dir):
    # Read the image
    image_path = os.path.join(image_dir, filename)
    image = cv2.imread(image_path)

    # Call the identify_person_with_unknown function to get the predicted identity and confidence
    predicted_identity, confidence = identify_person_with_unknown(image)

    # Add the results to the list
    results_list.append([filename, predicted_identity, confidence])

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Write the headers to the first row
worksheet.append(['filename', 'predicted_identity', 'confidence'])

# Write the results to the worksheet
for result in results_list:
    worksheet.append(result)

# Save the workbook to a file
workbook.save('/home/jawabreh/Desktop/ev/results-unknown-0.9.xlsx')


#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################


# Load face detection model
detector = MTCNN()

# Load face recognition model
facenet_model = load_model('/home/jawabreh/Desktop/face-recognition/facenet_keras.h5')

# Load face embeddings
data = np.load('/home/jawabreh/Desktop/face-recognition/embeddings/unmasked-embeddings.npz')
trainX, trainy = data['arr_0'], data['arr_1']

# Normalize input vectors
in_encoder = Normalizer(norm='l2')
trainX = in_encoder.transform(trainX)

# Label encode targets
out_encoder = LabelEncoder()
out_encoder.fit(trainy)
trainy = out_encoder.transform(trainy)

# Define the classes
class_names = out_encoder.classes_
class_names = np.append(class_names, 'unknown')

# Train SVM classifier
model = SVC(kernel='linear', probability=True)
model.fit(trainX, trainy)

# Define function to extract face embeddings
def extract_face_embeddings(image):
    # Detect faces in the image
    faces = detector.detect_faces(image)
    if not faces:
        return None
    # Extract the first face only
    x1, y1, width, height = faces[0]['box']
    x2, y2 = x1 + width, y1 + height
    face = image[y1:y2, x1:x2]
    # Resize face to the size required by facenet model
    face = cv2.resize(face, (160, 160))
    # Preprocess the face for facenet model
    face = face.astype('float32')
    mean, std = face.mean(), face.std()
    face = (face - mean) / std
    face = np.expand_dims(face, axis=0)
    # Generate embeddings using facenet model
    embeddings = facenet_model.predict(face)
    return embeddings[0]

# Define function to identify the identity of an input image
def identify_person(image):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    prediction = model.predict(embeddings)
    confidence = model.predict_proba(embeddings)[0][prediction] * 100
    prediction = out_encoder.inverse_transform(prediction)
    return prediction[0].item(), confidence

# Define function to identify the identity and confidence of an input image
def identify_person_with_unknown(image, threshold=0.99):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    predictions = model.predict_proba(embeddings)[0]
    max_idx = np.argmax(predictions)
    if predictions[max_idx] >= threshold:
        prediction = out_encoder.inverse_transform([max_idx])
        confidence = predictions[max_idx]
        return prediction[0].item(), confidence
    else:
        return None, None

# Define path to directory containing images
image_dir = '/home/jawabreh/Desktop/face-recognition/data/val/unknown/'


# Create an empty list to store the results
results_list = []

# Iterate through each image in the directory
for filename in os.listdir(image_dir):
    # Read the image
    image_path = os.path.join(image_dir, filename)
    image = cv2.imread(image_path)

    # Call the identify_person_with_unknown function to get the predicted identity and confidence
    predicted_identity, confidence = identify_person_with_unknown(image)

    # Add the results to the list
    results_list.append([filename, predicted_identity, confidence])

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Write the headers to the first row
worksheet.append(['filename', 'predicted_identity', 'confidence'])

# Write the results to the worksheet
for result in results_list:
    worksheet.append(result)

# Save the workbook to a file
workbook.save('/home/jawabreh/Desktop/ev/results-unknown-0.99.xlsx')


#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################


# Load face detection model
detector = MTCNN()

# Load face recognition model
facenet_model = load_model('/home/jawabreh/Desktop/face-recognition/facenet_keras.h5')

# Load face embeddings
data = np.load('/home/jawabreh/Desktop/face-recognition/embeddings/unmasked-embeddings.npz')
trainX, trainy = data['arr_0'], data['arr_1']

# Normalize input vectors
in_encoder = Normalizer(norm='l2')
trainX = in_encoder.transform(trainX)

# Label encode targets
out_encoder = LabelEncoder()
out_encoder.fit(trainy)
trainy = out_encoder.transform(trainy)

# Define the classes
class_names = out_encoder.classes_
class_names = np.append(class_names, 'unknown')

# Train SVM classifier
model = SVC(kernel='linear', probability=True)
model.fit(trainX, trainy)

# Define function to extract face embeddings
def extract_face_embeddings(image):
    # Detect faces in the image
    faces = detector.detect_faces(image)
    if not faces:
        return None
    # Extract the first face only
    x1, y1, width, height = faces[0]['box']
    x2, y2 = x1 + width, y1 + height
    face = image[y1:y2, x1:x2]
    # Resize face to the size required by facenet model
    face = cv2.resize(face, (160, 160))
    # Preprocess the face for facenet model
    face = face.astype('float32')
    mean, std = face.mean(), face.std()
    face = (face - mean) / std
    face = np.expand_dims(face, axis=0)
    # Generate embeddings using facenet model
    embeddings = facenet_model.predict(face)
    return embeddings[0]

# Define function to identify the identity of an input image
def identify_person(image):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    prediction = model.predict(embeddings)
    confidence = model.predict_proba(embeddings)[0][prediction] * 100
    prediction = out_encoder.inverse_transform(prediction)
    return prediction[0].item(), confidence

# Define function to identify the identity and confidence of an input image
def identify_person_with_unknown(image, threshold=0.85):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    predictions = model.predict_proba(embeddings)[0]
    max_idx = np.argmax(predictions)
    if predictions[max_idx] >= threshold:
        prediction = out_encoder.inverse_transform([max_idx])
        confidence = predictions[max_idx]
        return prediction[0].item(), confidence
    else:
        return None, None

# Define path to directory containing images
image_dir = '/home/jawabreh/Desktop/face-recognition/data/val/unknown/'


# Create an empty list to store the results
results_list = []

# Iterate through each image in the directory
for filename in os.listdir(image_dir):
    # Read the image
    image_path = os.path.join(image_dir, filename)
    image = cv2.imread(image_path)

    # Call the identify_person_with_unknown function to get the predicted identity and confidence
    predicted_identity, confidence = identify_person_with_unknown(image)

    # Add the results to the list
    results_list.append([filename, predicted_identity, confidence])

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Write the headers to the first row
worksheet.append(['filename', 'predicted_identity', 'confidence'])

# Write the results to the worksheet
for result in results_list:
    worksheet.append(result)

# Save the workbook to a file
workbook.save('/home/jawabreh/Desktop/ev/results-unknown-0.85.xlsx')

##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
############################################     NEW    ##################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################


# Load face detection model
detector = MTCNN()

# Load face recognition model
facenet_model = load_model('/home/jawabreh/Desktop/face-recognition/facenet_keras.h5')

# Load face embeddings
data = np.load('/home/jawabreh/Desktop/face-recognition/embeddings/unmasked-embeddings.npz')
trainX, trainy = data['arr_0'], data['arr_1']

# Normalize input vectors
in_encoder = Normalizer(norm='l2')
trainX = in_encoder.transform(trainX)

# Label encode targets
out_encoder = LabelEncoder()
out_encoder.fit(trainy)
trainy = out_encoder.transform(trainy)

# Define the classes
class_names = out_encoder.classes_
class_names = np.append(class_names, 'unknown')

# Train SVM classifier
model = SVC(kernel='linear', probability=True)
model.fit(trainX, trainy)

# Define function to extract face embeddings
def extract_face_embeddings(image):
    # Detect faces in the image
    faces = detector.detect_faces(image)
    if not faces:
        return None
    # Extract the first face only
    x1, y1, width, height = faces[0]['box']
    x2, y2 = x1 + width, y1 + height
    face = image[y1:y2, x1:x2]
    # Resize face to the size required by facenet model
    face = cv2.resize(face, (160, 160))
    # Preprocess the face for facenet model
    face = face.astype('float32')
    mean, std = face.mean(), face.std()
    face = (face - mean) / std
    face = np.expand_dims(face, axis=0)
    # Generate embeddings using facenet model
    embeddings = facenet_model.predict(face)
    return embeddings[0]

# Define function to identify the identity of an input image
def identify_person(image):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    prediction = model.predict(embeddings)
    confidence = model.predict_proba(embeddings)[0][prediction] * 100
    prediction = out_encoder.inverse_transform(prediction)
    return prediction[0].item(), confidence

# Define function to identify the identity and confidence of an input image
def identify_person_with_unknown(image, threshold=0.99):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    predictions = model.predict_proba(embeddings)[0]
    max_idx = np.argmax(predictions)
    if predictions[max_idx] >= threshold:
        prediction = out_encoder.inverse_transform([max_idx])
        confidence = predictions[max_idx]
        return prediction[0].item(), confidence
    else:
        return None, None

# Define path to directory containing images
image_dir = '/home/jawabreh/Desktop/face-recognition/data/val/ahmad/'


# Create an empty list to store the results
results_list = []

# Iterate through each image in the directory
for filename in os.listdir(image_dir):
    # Read the image
    image_path = os.path.join(image_dir, filename)
    image = cv2.imread(image_path)

    # Call the identify_person_with_unknown function to get the predicted identity and confidence
    predicted_identity, confidence = identify_person_with_unknown(image)

    # Add the results to the list
    results_list.append([filename, predicted_identity, confidence])

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Write the headers to the first row
worksheet.append(['filename', 'predicted_identity', 'confidence'])

# Write the results to the worksheet
for result in results_list:
    worksheet.append(result)

# Save the workbook to a file
workbook.save('/home/jawabreh/Desktop/ev/results-ahmad-0.99.xlsx')


#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################


# Load face detection model
detector = MTCNN()

# Load face recognition model
facenet_model = load_model('/home/jawabreh/Desktop/face-recognition/facenet_keras.h5')

# Load face embeddings
data = np.load('/home/jawabreh/Desktop/face-recognition/embeddings/unmasked-embeddings.npz')
trainX, trainy = data['arr_0'], data['arr_1']

# Normalize input vectors
in_encoder = Normalizer(norm='l2')
trainX = in_encoder.transform(trainX)

# Label encode targets
out_encoder = LabelEncoder()
out_encoder.fit(trainy)
trainy = out_encoder.transform(trainy)

# Define the classes
class_names = out_encoder.classes_
class_names = np.append(class_names, 'unknown')

# Train SVM classifier
model = SVC(kernel='linear', probability=True)
model.fit(trainX, trainy)

# Define function to extract face embeddings
def extract_face_embeddings(image):
    # Detect faces in the image
    faces = detector.detect_faces(image)
    if not faces:
        return None
    # Extract the first face only
    x1, y1, width, height = faces[0]['box']
    x2, y2 = x1 + width, y1 + height
    face = image[y1:y2, x1:x2]
    # Resize face to the size required by facenet model
    face = cv2.resize(face, (160, 160))
    # Preprocess the face for facenet model
    face = face.astype('float32')
    mean, std = face.mean(), face.std()
    face = (face - mean) / std
    face = np.expand_dims(face, axis=0)
    # Generate embeddings using facenet model
    embeddings = facenet_model.predict(face)
    return embeddings[0]

# Define function to identify the identity of an input image
def identify_person(image):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    prediction = model.predict(embeddings)
    confidence = model.predict_proba(embeddings)[0][prediction] * 100
    prediction = out_encoder.inverse_transform(prediction)
    return prediction[0].item(), confidence

# Define function to identify the identity and confidence of an input image
def identify_person_with_unknown(image, threshold=0.95):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    predictions = model.predict_proba(embeddings)[0]
    max_idx = np.argmax(predictions)
    if predictions[max_idx] >= threshold:
        prediction = out_encoder.inverse_transform([max_idx])
        confidence = predictions[max_idx]
        return prediction[0].item(), confidence
    else:
        return None, None

# Define path to directory containing images
image_dir = '/home/jawabreh/Desktop/face-recognition/data/val/ahmad/'


# Create an empty list to store the results
results_list = []

# Iterate through each image in the directory
for filename in os.listdir(image_dir):
    # Read the image
    image_path = os.path.join(image_dir, filename)
    image = cv2.imread(image_path)

    # Call the identify_person_with_unknown function to get the predicted identity and confidence
    predicted_identity, confidence = identify_person_with_unknown(image)

    # Add the results to the list
    results_list.append([filename, predicted_identity, confidence])

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Write the headers to the first row
worksheet.append(['filename', 'predicted_identity', 'confidence'])

# Write the results to the worksheet
for result in results_list:
    worksheet.append(result)

# Save the workbook to a file
workbook.save('/home/jawabreh/Desktop/ev/results-ahmad-0.95.xlsx')



#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################


# Load face detection model
detector = MTCNN()

# Load face recognition model
facenet_model = load_model('/home/jawabreh/Desktop/face-recognition/facenet_keras.h5')

# Load face embeddings
data = np.load('/home/jawabreh/Desktop/face-recognition/embeddings/unmasked-embeddings.npz')
trainX, trainy = data['arr_0'], data['arr_1']

# Normalize input vectors
in_encoder = Normalizer(norm='l2')
trainX = in_encoder.transform(trainX)

# Label encode targets
out_encoder = LabelEncoder()
out_encoder.fit(trainy)
trainy = out_encoder.transform(trainy)

# Define the classes
class_names = out_encoder.classes_
class_names = np.append(class_names, 'unknown')

# Train SVM classifier
model = SVC(kernel='linear', probability=True)
model.fit(trainX, trainy)

# Define function to extract face embeddings
def extract_face_embeddings(image):
    # Detect faces in the image
    faces = detector.detect_faces(image)
    if not faces:
        return None
    # Extract the first face only
    x1, y1, width, height = faces[0]['box']
    x2, y2 = x1 + width, y1 + height
    face = image[y1:y2, x1:x2]
    # Resize face to the size required by facenet model
    face = cv2.resize(face, (160, 160))
    # Preprocess the face for facenet model
    face = face.astype('float32')
    mean, std = face.mean(), face.std()
    face = (face - mean) / std
    face = np.expand_dims(face, axis=0)
    # Generate embeddings using facenet model
    embeddings = facenet_model.predict(face)
    return embeddings[0]

# Define function to identify the identity of an input image
def identify_person(image):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    prediction = model.predict(embeddings)
    confidence = model.predict_proba(embeddings)[0][prediction] * 100
    prediction = out_encoder.inverse_transform(prediction)
    return prediction[0].item(), confidence

# Define function to identify the identity and confidence of an input image
def identify_person_with_unknown(image, threshold=0.90):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    predictions = model.predict_proba(embeddings)[0]
    max_idx = np.argmax(predictions)
    if predictions[max_idx] >= threshold:
        prediction = out_encoder.inverse_transform([max_idx])
        confidence = predictions[max_idx]
        return prediction[0].item(), confidence
    else:
        return None, None

# Define path to directory containing images
image_dir = '/home/jawabreh/Desktop/face-recognition/data/val/ahmad/'


# Create an empty list to store the results
results_list = []

# Iterate through each image in the directory
for filename in os.listdir(image_dir):
    # Read the image
    image_path = os.path.join(image_dir, filename)
    image = cv2.imread(image_path)

    # Call the identify_person_with_unknown function to get the predicted identity and confidence
    predicted_identity, confidence = identify_person_with_unknown(image)

    # Add the results to the list
    results_list.append([filename, predicted_identity, confidence])

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Write the headers to the first row
worksheet.append(['filename', 'predicted_identity', 'confidence'])

# Write the results to the worksheet
for result in results_list:
    worksheet.append(result)

# Save the workbook to a file
workbook.save('/home/jawabreh/Desktop/ev/results-ahmad-0.9.xlsx')


#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################


# Load face detection model
detector = MTCNN()

# Load face recognition model
facenet_model = load_model('/home/jawabreh/Desktop/face-recognition/facenet_keras.h5')

# Load face embeddings
data = np.load('/home/jawabreh/Desktop/face-recognition/embeddings/unmasked-embeddings.npz')
trainX, trainy = data['arr_0'], data['arr_1']

# Normalize input vectors
in_encoder = Normalizer(norm='l2')
trainX = in_encoder.transform(trainX)

# Label encode targets
out_encoder = LabelEncoder()
out_encoder.fit(trainy)
trainy = out_encoder.transform(trainy)

# Define the classes
class_names = out_encoder.classes_
class_names = np.append(class_names, 'unknown')

# Train SVM classifier
model = SVC(kernel='linear', probability=True)
model.fit(trainX, trainy)

# Define function to extract face embeddings
def extract_face_embeddings(image):
    # Detect faces in the image
    faces = detector.detect_faces(image)
    if not faces:
        return None
    # Extract the first face only
    x1, y1, width, height = faces[0]['box']
    x2, y2 = x1 + width, y1 + height
    face = image[y1:y2, x1:x2]
    # Resize face to the size required by facenet model
    face = cv2.resize(face, (160, 160))
    # Preprocess the face for facenet model
    face = face.astype('float32')
    mean, std = face.mean(), face.std()
    face = (face - mean) / std
    face = np.expand_dims(face, axis=0)
    # Generate embeddings using facenet model
    embeddings = facenet_model.predict(face)
    return embeddings[0]

# Define function to identify the identity of an input image
def identify_person(image):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    prediction = model.predict(embeddings)
    confidence = model.predict_proba(embeddings)[0][prediction] * 100
    prediction = out_encoder.inverse_transform(prediction)
    return prediction[0].item(), confidence

# Define function to identify the identity and confidence of an input image
def identify_person_with_unknown(image, threshold=0.85):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    predictions = model.predict_proba(embeddings)[0]
    max_idx = np.argmax(predictions)
    if predictions[max_idx] >= threshold:
        prediction = out_encoder.inverse_transform([max_idx])
        confidence = predictions[max_idx]
        return prediction[0].item(), confidence
    else:
        return None, None

# Define path to directory containing images
image_dir = '/home/jawabreh/Desktop/face-recognition/data/val/ahmad/'


# Create an empty list to store the results
results_list = []

# Iterate through each image in the directory
for filename in os.listdir(image_dir):
    # Read the image
    image_path = os.path.join(image_dir, filename)
    image = cv2.imread(image_path)

    # Call the identify_person_with_unknown function to get the predicted identity and confidence
    predicted_identity, confidence = identify_person_with_unknown(image)

    # Add the results to the list
    results_list.append([filename, predicted_identity, confidence])

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Write the headers to the first row
worksheet.append(['filename', 'predicted_identity', 'confidence'])

# Write the results to the worksheet
for result in results_list:
    worksheet.append(result)

# Save the workbook to a file
workbook.save('/home/jawabreh/Desktop/ev/results-ahmad-0.85.xlsx')


##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
############################################     NEW    ##################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################

# Load face detection model
detector = MTCNN()

# Load face recognition model
facenet_model = load_model('/home/jawabreh/Desktop/face-recognition/facenet_keras.h5')

# Load face embeddings
data = np.load('/home/jawabreh/Desktop/face-recognition/embeddings/unmasked-embeddings.npz')
trainX, trainy = data['arr_0'], data['arr_1']

# Normalize input vectors
in_encoder = Normalizer(norm='l2')
trainX = in_encoder.transform(trainX)

# Label encode targets
out_encoder = LabelEncoder()
out_encoder.fit(trainy)
trainy = out_encoder.transform(trainy)

# Define the classes
class_names = out_encoder.classes_
class_names = np.append(class_names, 'unknown')

# Train SVM classifier
model = SVC(kernel='linear', probability=True)
model.fit(trainX, trainy)

# Define function to extract face embeddings
def extract_face_embeddings(image):
    # Detect faces in the image
    faces = detector.detect_faces(image)
    if not faces:
        return None
    # Extract the first face only
    x1, y1, width, height = faces[0]['box']
    x2, y2 = x1 + width, y1 + height
    face = image[y1:y2, x1:x2]
    # Resize face to the size required by facenet model
    face = cv2.resize(face, (160, 160))
    # Preprocess the face for facenet model
    face = face.astype('float32')
    mean, std = face.mean(), face.std()
    face = (face - mean) / std
    face = np.expand_dims(face, axis=0)
    # Generate embeddings using facenet model
    embeddings = facenet_model.predict(face)
    return embeddings[0]

# Define function to identify the identity of an input image
def identify_person(image):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    prediction = model.predict(embeddings)
    confidence = model.predict_proba(embeddings)[0][prediction] * 100
    prediction = out_encoder.inverse_transform(prediction)
    return prediction[0].item(), confidence

# Define function to identify the identity and confidence of an input image
def identify_person_with_unknown(image, threshold=0.99):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    predictions = model.predict_proba(embeddings)[0]
    max_idx = np.argmax(predictions)
    if predictions[max_idx] >= threshold:
        prediction = out_encoder.inverse_transform([max_idx])
        confidence = predictions[max_idx]
        return prediction[0].item(), confidence
    else:
        return None, None

# Define path to directory containing images
image_dir = '/home/jawabreh/Desktop/face-recognition/data/val/milena/'


# Create an empty list to store the results
results_list = []

# Iterate through each image in the directory
for filename in os.listdir(image_dir):
    # Read the image
    image_path = os.path.join(image_dir, filename)
    image = cv2.imread(image_path)

    # Call the identify_person_with_unknown function to get the predicted identity and confidence
    predicted_identity, confidence = identify_person_with_unknown(image)

    # Add the results to the list
    results_list.append([filename, predicted_identity, confidence])

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Write the headers to the first row
worksheet.append(['filename', 'predicted_identity', 'confidence'])

# Write the results to the worksheet
for result in results_list:
    worksheet.append(result)

# Save the workbook to a file
workbook.save('/home/jawabreh/Desktop/ev/results-milena-0.99.xlsx')


#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################


# Load face detection model
detector = MTCNN()

# Load face recognition model
facenet_model = load_model('/home/jawabreh/Desktop/face-recognition/facenet_keras.h5')

# Load face embeddings
data = np.load('/home/jawabreh/Desktop/face-recognition/embeddings/unmasked-embeddings.npz')
trainX, trainy = data['arr_0'], data['arr_1']

# Normalize input vectors
in_encoder = Normalizer(norm='l2')
trainX = in_encoder.transform(trainX)

# Label encode targets
out_encoder = LabelEncoder()
out_encoder.fit(trainy)
trainy = out_encoder.transform(trainy)

# Define the classes
class_names = out_encoder.classes_
class_names = np.append(class_names, 'unknown')

# Train SVM classifier
model = SVC(kernel='linear', probability=True)
model.fit(trainX, trainy)

# Define function to extract face embeddings
def extract_face_embeddings(image):
    # Detect faces in the image
    faces = detector.detect_faces(image)
    if not faces:
        return None
    # Extract the first face only
    x1, y1, width, height = faces[0]['box']
    x2, y2 = x1 + width, y1 + height
    face = image[y1:y2, x1:x2]
    # Resize face to the size required by facenet model
    face = cv2.resize(face, (160, 160))
    # Preprocess the face for facenet model
    face = face.astype('float32')
    mean, std = face.mean(), face.std()
    face = (face - mean) / std
    face = np.expand_dims(face, axis=0)
    # Generate embeddings using facenet model
    embeddings = facenet_model.predict(face)
    return embeddings[0]

# Define function to identify the identity of an input image
def identify_person(image):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    prediction = model.predict(embeddings)
    confidence = model.predict_proba(embeddings)[0][prediction] * 100
    prediction = out_encoder.inverse_transform(prediction)
    return prediction[0].item(), confidence

# Define function to identify the identity and confidence of an input image
def identify_person_with_unknown(image, threshold=0.95):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    predictions = model.predict_proba(embeddings)[0]
    max_idx = np.argmax(predictions)
    if predictions[max_idx] >= threshold:
        prediction = out_encoder.inverse_transform([max_idx])
        confidence = predictions[max_idx]
        return prediction[0].item(), confidence
    else:
        return None, None

# Define path to directory containing images
image_dir = '/home/jawabreh/Desktop/face-recognition/data/val/milena/'


# Create an empty list to store the results
results_list = []

# Iterate through each image in the directory
for filename in os.listdir(image_dir):
    # Read the image
    image_path = os.path.join(image_dir, filename)
    image = cv2.imread(image_path)

    # Call the identify_person_with_unknown function to get the predicted identity and confidence
    predicted_identity, confidence = identify_person_with_unknown(image)

    # Add the results to the list
    results_list.append([filename, predicted_identity, confidence])

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Write the headers to the first row
worksheet.append(['filename', 'predicted_identity', 'confidence'])

# Write the results to the worksheet
for result in results_list:
    worksheet.append(result)

# Save the workbook to a file
workbook.save('/home/jawabreh/Desktop/ev/results-milena-0.95.xlsx')



#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################


# Load face detection model
detector = MTCNN()

# Load face recognition model
facenet_model = load_model('/home/jawabreh/Desktop/face-recognition/facenet_keras.h5')

# Load face embeddings
data = np.load('/home/jawabreh/Desktop/face-recognition/embeddings/unmasked-embeddings.npz')
trainX, trainy = data['arr_0'], data['arr_1']

# Normalize input vectors
in_encoder = Normalizer(norm='l2')
trainX = in_encoder.transform(trainX)

# Label encode targets
out_encoder = LabelEncoder()
out_encoder.fit(trainy)
trainy = out_encoder.transform(trainy)

# Define the classes
class_names = out_encoder.classes_
class_names = np.append(class_names, 'unknown')

# Train SVM classifier
model = SVC(kernel='linear', probability=True)
model.fit(trainX, trainy)

# Define function to extract face embeddings
def extract_face_embeddings(image):
    # Detect faces in the image
    faces = detector.detect_faces(image)
    if not faces:
        return None
    # Extract the first face only
    x1, y1, width, height = faces[0]['box']
    x2, y2 = x1 + width, y1 + height
    face = image[y1:y2, x1:x2]
    # Resize face to the size required by facenet model
    face = cv2.resize(face, (160, 160))
    # Preprocess the face for facenet model
    face = face.astype('float32')
    mean, std = face.mean(), face.std()
    face = (face - mean) / std
    face = np.expand_dims(face, axis=0)
    # Generate embeddings using facenet model
    embeddings = facenet_model.predict(face)
    return embeddings[0]

# Define function to identify the identity of an input image
def identify_person(image):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    prediction = model.predict(embeddings)
    confidence = model.predict_proba(embeddings)[0][prediction] * 100
    prediction = out_encoder.inverse_transform(prediction)
    return prediction[0].item(), confidence

# Define function to identify the identity and confidence of an input image
def identify_person_with_unknown(image, threshold=0.90):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    predictions = model.predict_proba(embeddings)[0]
    max_idx = np.argmax(predictions)
    if predictions[max_idx] >= threshold:
        prediction = out_encoder.inverse_transform([max_idx])
        confidence = predictions[max_idx]
        return prediction[0].item(), confidence
    else:
        return None, None

# Define path to directory containing images
image_dir = '/home/jawabreh/Desktop/face-recognition/data/val/milena/'


# Create an empty list to store the results
results_list = []

# Iterate through each image in the directory
for filename in os.listdir(image_dir):
    # Read the image
    image_path = os.path.join(image_dir, filename)
    image = cv2.imread(image_path)

    # Call the identify_person_with_unknown function to get the predicted identity and confidence
    predicted_identity, confidence = identify_person_with_unknown(image)

    # Add the results to the list
    results_list.append([filename, predicted_identity, confidence])

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Write the headers to the first row
worksheet.append(['filename', 'predicted_identity', 'confidence'])

# Write the results to the worksheet
for result in results_list:
    worksheet.append(result)

# Save the workbook to a file
workbook.save('/home/jawabreh/Desktop/ev/results-milena-0.9.xlsx')


#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################


# Load face detection model
detector = MTCNN()

# Load face recognition model
facenet_model = load_model('/home/jawabreh/Desktop/face-recognition/facenet_keras.h5')

# Load face embeddings
data = np.load('/home/jawabreh/Desktop/face-recognition/embeddings/unmasked-embeddings.npz')
trainX, trainy = data['arr_0'], data['arr_1']

# Normalize input vectors
in_encoder = Normalizer(norm='l2')
trainX = in_encoder.transform(trainX)

# Label encode targets
out_encoder = LabelEncoder()
out_encoder.fit(trainy)
trainy = out_encoder.transform(trainy)

# Define the classes
class_names = out_encoder.classes_
class_names = np.append(class_names, 'unknown')

# Train SVM classifier
model = SVC(kernel='linear', probability=True)
model.fit(trainX, trainy)

# Define function to extract face embeddings
def extract_face_embeddings(image):
    # Detect faces in the image
    faces = detector.detect_faces(image)
    if not faces:
        return None
    # Extract the first face only
    x1, y1, width, height = faces[0]['box']
    x2, y2 = x1 + width, y1 + height
    face = image[y1:y2, x1:x2]
    # Resize face to the size required by facenet model
    face = cv2.resize(face, (160, 160))
    # Preprocess the face for facenet model
    face = face.astype('float32')
    mean, std = face.mean(), face.std()
    face = (face - mean) / std
    face = np.expand_dims(face, axis=0)
    # Generate embeddings using facenet model
    embeddings = facenet_model.predict(face)
    return embeddings[0]

# Define function to identify the identity of an input image
def identify_person(image):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    prediction = model.predict(embeddings)
    confidence = model.predict_proba(embeddings)[0][prediction] * 100
    prediction = out_encoder.inverse_transform(prediction)
    return prediction[0].item(), confidence

# Define function to identify the identity and confidence of an input image
def identify_person_with_unknown(image, threshold=0.85):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    predictions = model.predict_proba(embeddings)[0]
    max_idx = np.argmax(predictions)
    if predictions[max_idx] >= threshold:
        prediction = out_encoder.inverse_transform([max_idx])
        confidence = predictions[max_idx]
        return prediction[0].item(), confidence
    else:
        return None, None

# Define path to directory containing images
image_dir = '/home/jawabreh/Desktop/face-recognition/data/val/milena/'


# Create an empty list to store the results
results_list = []

# Iterate through each image in the directory
for filename in os.listdir(image_dir):
    # Read the image
    image_path = os.path.join(image_dir, filename)
    image = cv2.imread(image_path)

    # Call the identify_person_with_unknown function to get the predicted identity and confidence
    predicted_identity, confidence = identify_person_with_unknown(image)

    # Add the results to the list
    results_list.append([filename, predicted_identity, confidence])

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Write the headers to the first row
worksheet.append(['filename', 'predicted_identity', 'confidence'])

# Write the results to the worksheet
for result in results_list:
    worksheet.append(result)

# Save the workbook to a file
workbook.save('/home/jawabreh/Desktop/ev/results-milena-0.85.xlsx')
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
############################################     NEW    ##################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################
##########################################################################################################################################

# Load face detection model
detector = MTCNN()

# Load face recognition model
facenet_model = load_model('/home/jawabreh/Desktop/face-recognition/facenet_keras.h5')

# Load face embeddings
data = np.load('/home/jawabreh/Desktop/face-recognition/embeddings/unmasked-embeddings.npz')
trainX, trainy = data['arr_0'], data['arr_1']

# Normalize input vectors
in_encoder = Normalizer(norm='l2')
trainX = in_encoder.transform(trainX)

# Label encode targets
out_encoder = LabelEncoder()
out_encoder.fit(trainy)
trainy = out_encoder.transform(trainy)

# Define the classes
class_names = out_encoder.classes_
class_names = np.append(class_names, 'unknown')

# Train SVM classifier
model = SVC(kernel='linear', probability=True)
model.fit(trainX, trainy)

# Define function to extract face embeddings
def extract_face_embeddings(image):
    # Detect faces in the image
    faces = detector.detect_faces(image)
    if not faces:
        return None
    # Extract the first face only
    x1, y1, width, height = faces[0]['box']
    x2, y2 = x1 + width, y1 + height
    face = image[y1:y2, x1:x2]
    # Resize face to the size required by facenet model
    face = cv2.resize(face, (160, 160))
    # Preprocess the face for facenet model
    face = face.astype('float32')
    mean, std = face.mean(), face.std()
    face = (face - mean) / std
    face = np.expand_dims(face, axis=0)
    # Generate embeddings using facenet model
    embeddings = facenet_model.predict(face)
    return embeddings[0]

# Define function to identify the identity of an input image
def identify_person(image):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    prediction = model.predict(embeddings)
    confidence = model.predict_proba(embeddings)[0][prediction] * 100
    prediction = out_encoder.inverse_transform(prediction)
    return prediction[0].item(), confidence

# Define function to identify the identity and confidence of an input image
def identify_person_with_unknown(image, threshold=0.99):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    predictions = model.predict_proba(embeddings)[0]
    max_idx = np.argmax(predictions)
    if predictions[max_idx] >= threshold:
        prediction = out_encoder.inverse_transform([max_idx])
        confidence = predictions[max_idx]
        return prediction[0].item(), confidence
    else:
        return None, None

# Define path to directory containing images
image_dir = '/home/jawabreh/Desktop/face-recognition/data/val/ravilya/'


# Create an empty list to store the results
results_list = []

# Iterate through each image in the directory
for filename in os.listdir(image_dir):
    # Read the image
    image_path = os.path.join(image_dir, filename)
    image = cv2.imread(image_path)

    # Call the identify_person_with_unknown function to get the predicted identity and confidence
    predicted_identity, confidence = identify_person_with_unknown(image)

    # Add the results to the list
    results_list.append([filename, predicted_identity, confidence])

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Write the headers to the first row
worksheet.append(['filename', 'predicted_identity', 'confidence'])

# Write the results to the worksheet
for result in results_list:
    worksheet.append(result)

# Save the workbook to a file
workbook.save('/home/jawabreh/Desktop/ev/results-ravilya-0.99.xlsx')


#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################


# Load face detection model
detector = MTCNN()

# Load face recognition model
facenet_model = load_model('/home/jawabreh/Desktop/face-recognition/facenet_keras.h5')

# Load face embeddings
data = np.load('/home/jawabreh/Desktop/face-recognition/embeddings/unmasked-embeddings.npz')
trainX, trainy = data['arr_0'], data['arr_1']

# Normalize input vectors
in_encoder = Normalizer(norm='l2')
trainX = in_encoder.transform(trainX)

# Label encode targets
out_encoder = LabelEncoder()
out_encoder.fit(trainy)
trainy = out_encoder.transform(trainy)

# Define the classes
class_names = out_encoder.classes_
class_names = np.append(class_names, 'unknown')

# Train SVM classifier
model = SVC(kernel='linear', probability=True)
model.fit(trainX, trainy)

# Define function to extract face embeddings
def extract_face_embeddings(image):
    # Detect faces in the image
    faces = detector.detect_faces(image)
    if not faces:
        return None
    # Extract the first face only
    x1, y1, width, height = faces[0]['box']
    x2, y2 = x1 + width, y1 + height
    face = image[y1:y2, x1:x2]
    # Resize face to the size required by facenet model
    face = cv2.resize(face, (160, 160))
    # Preprocess the face for facenet model
    face = face.astype('float32')
    mean, std = face.mean(), face.std()
    face = (face - mean) / std
    face = np.expand_dims(face, axis=0)
    # Generate embeddings using facenet model
    embeddings = facenet_model.predict(face)
    return embeddings[0]

# Define function to identify the identity of an input image
def identify_person(image):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    prediction = model.predict(embeddings)
    confidence = model.predict_proba(embeddings)[0][prediction] * 100
    prediction = out_encoder.inverse_transform(prediction)
    return prediction[0].item(), confidence

# Define function to identify the identity and confidence of an input image
def identify_person_with_unknown(image, threshold=0.95):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    predictions = model.predict_proba(embeddings)[0]
    max_idx = np.argmax(predictions)
    if predictions[max_idx] >= threshold:
        prediction = out_encoder.inverse_transform([max_idx])
        confidence = predictions[max_idx]
        return prediction[0].item(), confidence
    else:
        return None, None

# Define path to directory containing images
image_dir = '/home/jawabreh/Desktop/face-recognition/data/val/ravilya/'


# Create an empty list to store the results
results_list = []

# Iterate through each image in the directory
for filename in os.listdir(image_dir):
    # Read the image
    image_path = os.path.join(image_dir, filename)
    image = cv2.imread(image_path)

    # Call the identify_person_with_unknown function to get the predicted identity and confidence
    predicted_identity, confidence = identify_person_with_unknown(image)

    # Add the results to the list
    results_list.append([filename, predicted_identity, confidence])

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Write the headers to the first row
worksheet.append(['filename', 'predicted_identity', 'confidence'])

# Write the results to the worksheet
for result in results_list:
    worksheet.append(result)

# Save the workbook to a file
workbook.save('/home/jawabreh/Desktop/ev/results-ravilya-0.95.xlsx')



#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################


# Load face detection model
detector = MTCNN()

# Load face recognition model
facenet_model = load_model('/home/jawabreh/Desktop/face-recognition/facenet_keras.h5')

# Load face embeddings
data = np.load('/home/jawabreh/Desktop/face-recognition/embeddings/unmasked-embeddings.npz')
trainX, trainy = data['arr_0'], data['arr_1']

# Normalize input vectors
in_encoder = Normalizer(norm='l2')
trainX = in_encoder.transform(trainX)

# Label encode targets
out_encoder = LabelEncoder()
out_encoder.fit(trainy)
trainy = out_encoder.transform(trainy)

# Define the classes
class_names = out_encoder.classes_
class_names = np.append(class_names, 'unknown')

# Train SVM classifier
model = SVC(kernel='linear', probability=True)
model.fit(trainX, trainy)

# Define function to extract face embeddings
def extract_face_embeddings(image):
    # Detect faces in the image
    faces = detector.detect_faces(image)
    if not faces:
        return None
    # Extract the first face only
    x1, y1, width, height = faces[0]['box']
    x2, y2 = x1 + width, y1 + height
    face = image[y1:y2, x1:x2]
    # Resize face to the size required by facenet model
    face = cv2.resize(face, (160, 160))
    # Preprocess the face for facenet model
    face = face.astype('float32')
    mean, std = face.mean(), face.std()
    face = (face - mean) / std
    face = np.expand_dims(face, axis=0)
    # Generate embeddings using facenet model
    embeddings = facenet_model.predict(face)
    return embeddings[0]

# Define function to identify the identity of an input image
def identify_person(image):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    prediction = model.predict(embeddings)
    confidence = model.predict_proba(embeddings)[0][prediction] * 100
    prediction = out_encoder.inverse_transform(prediction)
    return prediction[0].item(), confidence

# Define function to identify the identity and confidence of an input image
def identify_person_with_unknown(image, threshold=0.90):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    predictions = model.predict_proba(embeddings)[0]
    max_idx = np.argmax(predictions)
    if predictions[max_idx] >= threshold:
        prediction = out_encoder.inverse_transform([max_idx])
        confidence = predictions[max_idx]
        return prediction[0].item(), confidence
    else:
        return None, None

# Define path to directory containing images
image_dir = '/home/jawabreh/Desktop/face-recognition/data/val/ravilya/'


# Create an empty list to store the results
results_list = []

# Iterate through each image in the directory
for filename in os.listdir(image_dir):
    # Read the image
    image_path = os.path.join(image_dir, filename)
    image = cv2.imread(image_path)

    # Call the identify_person_with_unknown function to get the predicted identity and confidence
    predicted_identity, confidence = identify_person_with_unknown(image)

    # Add the results to the list
    results_list.append([filename, predicted_identity, confidence])

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Write the headers to the first row
worksheet.append(['filename', 'predicted_identity', 'confidence'])

# Write the results to the worksheet
for result in results_list:
    worksheet.append(result)

# Save the workbook to a file
workbook.save('/home/jawabreh/Desktop/ev/results-ravilya-0.9.xlsx')


#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################
#######################################################


# Load face detection model
detector = MTCNN()

# Load face recognition model
facenet_model = load_model('/home/jawabreh/Desktop/face-recognition/facenet_keras.h5')

# Load face embeddings
data = np.load('/home/jawabreh/Desktop/face-recognition/embeddings/unmasked-embeddings.npz')
trainX, trainy = data['arr_0'], data['arr_1']

# Normalize input vectors
in_encoder = Normalizer(norm='l2')
trainX = in_encoder.transform(trainX)

# Label encode targets
out_encoder = LabelEncoder()
out_encoder.fit(trainy)
trainy = out_encoder.transform(trainy)

# Define the classes
class_names = out_encoder.classes_
class_names = np.append(class_names, 'unknown')

# Train SVM classifier
model = SVC(kernel='linear', probability=True)
model.fit(trainX, trainy)

# Define function to extract face embeddings
def extract_face_embeddings(image):
    # Detect faces in the image
    faces = detector.detect_faces(image)
    if not faces:
        return None
    # Extract the first face only
    x1, y1, width, height = faces[0]['box']
    x2, y2 = x1 + width, y1 + height
    face = image[y1:y2, x1:x2]
    # Resize face to the size required by facenet model
    face = cv2.resize(face, (160, 160))
    # Preprocess the face for facenet model
    face = face.astype('float32')
    mean, std = face.mean(), face.std()
    face = (face - mean) / std
    face = np.expand_dims(face, axis=0)
    # Generate embeddings using facenet model
    embeddings = facenet_model.predict(face)
    return embeddings[0]

# Define function to identify the identity of an input image
def identify_person(image):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    prediction = model.predict(embeddings)
    confidence = model.predict_proba(embeddings)[0][prediction] * 100
    prediction = out_encoder.inverse_transform(prediction)
    return prediction[0].item(), confidence

# Define function to identify the identity and confidence of an input image
def identify_person_with_unknown(image, threshold=0.85):
    # Extract face embeddings from input image
    embeddings = extract_face_embeddings(image)
    if embeddings is None:
        return None, None
    # Normalize embeddings
    embeddings = in_encoder.transform([embeddings])
    # Predict the identity and confidence using SVM classifier
    predictions = model.predict_proba(embeddings)[0]
    max_idx = np.argmax(predictions)
    if predictions[max_idx] >= threshold:
        prediction = out_encoder.inverse_transform([max_idx])
        confidence = predictions[max_idx]
        return prediction[0].item(), confidence
    else:
        return None, None

# Define path to directory containing images
image_dir = '/home/jawabreh/Desktop/face-recognition/data/val/ravilya/'


# Create an empty list to store the results
results_list = []

# Iterate through each image in the directory
for filename in os.listdir(image_dir):
    # Read the image
    image_path = os.path.join(image_dir, filename)
    image = cv2.imread(image_path)

    # Call the identify_person_with_unknown function to get the predicted identity and confidence
    predicted_identity, confidence = identify_person_with_unknown(image)

    # Add the results to the list
    results_list.append([filename, predicted_identity, confidence])

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Write the headers to the first row
worksheet.append(['filename', 'predicted_identity', 'confidence'])

# Write the results to the worksheet
for result in results_list:
    worksheet.append(result)

# Save the workbook to a file
workbook.save('/home/jawabreh/Desktop/ev/results-ravilya-0.85.xlsx')